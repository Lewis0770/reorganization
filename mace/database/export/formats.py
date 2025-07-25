"""
Multi-format Export for MACE Database
=====================================
Supports exporting materials data to various formats:
- CSV (default)
- JSON
- Excel (.xlsx)
- LaTeX tables
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime


class ExportFormatter:
    """Handles formatting and exporting of materials data to various formats."""
    
    def __init__(self):
        """Initialize the formatter."""
        self.supported_formats = ['csv', 'json', 'excel', 'latex', 'html']
        
    def export(self, data: List[Dict], output_file: str, format: str = 'csv',
               properties_filter: List[str] = None,
               include_structures: bool = False,
               metadata: Dict = None) -> bool:
        """
        Export data to specified format.
        
        Args:
            data: List of records to export (materials or properties)
            output_file: Output file path
            format: Export format (csv, json, excel, latex, html)
            properties_filter: List of property names to include (None = all)
            include_structures: Whether to include structure data
            metadata: Additional metadata to include in export
            
        Returns:
            True if successful
        """
        format = format.lower()
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}. Choose from {self.supported_formats}")
            
        # Filter data if needed
        if properties_filter and data:
            data = self._filter_properties(data, properties_filter)
            
        # Remove structure data if not requested
        if not include_structures and data:
            data = self._remove_structure_data(data)
            
        # Export based on format
        if format == 'csv':
            return self._export_csv(data, output_file)
        elif format == 'json':
            return self._export_json(data, output_file, metadata)
        elif format == 'excel':
            return self._export_excel(data, output_file, metadata)
        elif format == 'latex':
            return self._export_latex(data, output_file, metadata)
        elif format == 'html':
            return self._export_html(data, output_file, metadata)
            
    def _filter_properties(self, data: List[Dict], properties: List[str]) -> List[Dict]:
        """Filter data to include only specified properties."""
        filtered = []
        for record in data:
            filtered_record = {}
            for key, value in record.items():
                if key in properties or key in ['material_id', 'calc_id']:  # Always include IDs
                    filtered_record[key] = value
            filtered.append(filtered_record)
        return filtered
        
    def _remove_structure_data(self, data: List[Dict]) -> List[Dict]:
        """Remove large structure-related fields."""
        structure_fields = ['structure_json', 'structure_ase', 'atomic_positions',
                          'initial_atomic_positions', 'final_atomic_positions']
        
        cleaned = []
        for record in data:
            cleaned_record = {}
            for key, value in record.items():
                if key not in structure_fields:
                    cleaned_record[key] = value
            cleaned.append(cleaned_record)
        return cleaned
        
    def _export_csv(self, data: List[Dict], output_file: str) -> bool:
        """Export to CSV format."""
        if not data:
            return False
            
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Get all unique keys
        all_keys = set()
        for record in data:
            all_keys.update(record.keys())
        fieldnames = sorted(all_keys)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
            
        return True
        
    def _export_json(self, data: List[Dict], output_file: str, metadata: Dict = None) -> bool:
        """Export to JSON format."""
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'record_count': len(data),
                'mace_version': '1.0.0'
            },
            'data': data
        }
        
        if metadata:
            export_data['metadata'].update(metadata)
            
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str)
            
        return True
        
    def _export_excel(self, data: List[Dict], output_file: str, metadata: Dict = None) -> bool:
        """Export to Excel format with formatting."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("Excel export requires pandas and openpyxl: pip install pandas openpyxl")
            
        if not data:
            return False
            
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Materials Data"
        
        # Add metadata sheet if provided
        if metadata:
            meta_ws = wb.create_sheet("Metadata")
            meta_ws['A1'] = "Export Information"
            meta_ws['A1'].font = Font(bold=True, size=14)
            
            row = 3
            for key, value in metadata.items():
                meta_ws[f'A{row}'] = key
                meta_ws[f'B{row}'] = str(value)
                row += 1
                
            meta_ws['A{}'.format(row+1)] = "Export Date"
            meta_ws['B{}'.format(row+1)] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
        # Write data to main sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        # Format header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Save workbook
        wb.save(output_path)
        return True
        
    def _export_latex(self, data: List[Dict], output_file: str, metadata: Dict = None) -> bool:
        """Export to LaTeX table format."""
        if not data:
            return False
            
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Select key columns for LaTeX table (too many columns don't work well)
        key_columns = ['material_id', 'formula', 'space_group', 'band_gap', 
                      'total_energy', 'optimization_converged']
        
        # Filter to available columns
        available_columns = []
        if data:
            first_record = data[0]
            for col in key_columns:
                if col in first_record:
                    available_columns.append(col)
                    
        if not available_columns:
            # Fallback to first 6 columns
            available_columns = list(data[0].keys())[:6]
            
        with open(output_path, 'w', encoding='utf-8') as f:
            # Write LaTeX preamble
            f.write("% MACE Database Export\n")
            f.write(f"% Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            if metadata:
                f.write(f"% Metadata: {metadata}\n")
            f.write("\n")
            
            # Begin table
            f.write("\\begin{table}[htbp]\n")
            f.write("\\centering\n")
            f.write("\\caption{Materials Database Export}\n")
            f.write("\\label{tab:materials}\n")
            
            # Column specification
            col_spec = "|" + "|".join(["c"] * len(available_columns)) + "|"
            f.write(f"\\begin{{tabular}}{{{col_spec}}}\n")
            f.write("\\hline\n")
            
            # Header row
            headers = [self._latex_escape(col.replace('_', ' ').title()) 
                      for col in available_columns]
            f.write(" & ".join(headers) + " \\\\\n")
            f.write("\\hline\\hline\n")
            
            # Data rows
            for record in data[:50]:  # Limit to 50 rows for LaTeX
                row_data = []
                for col in available_columns:
                    value = record.get(col, '')
                    # Format numeric values
                    if isinstance(value, float):
                        value = f"{value:.4f}"
                    row_data.append(self._latex_escape(str(value)))
                f.write(" & ".join(row_data) + " \\\\\n")
                f.write("\\hline\n")
                
            # End table
            f.write("\\end{tabular}\n")
            f.write("\\end{table}\n")
            
            if len(data) > 50:
                f.write(f"\n% Note: Showing first 50 of {len(data)} total records\n")
                
        return True
        
    def _export_html(self, data: List[Dict], output_file: str, metadata: Dict = None) -> bool:
        """Export to HTML table format with styling."""
        if not data:
            return False
            
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            # HTML header
            f.write("""<!DOCTYPE html>
<html>
<head>
    <title>MACE Database Export</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { color: #366092; }
        table { border-collapse: collapse; width: 100%; margin-top: 20px; }
        th { background-color: #366092; color: white; padding: 10px; text-align: left; }
        td { border: 1px solid #ddd; padding: 8px; }
        tr:nth-child(even) { background-color: #f2f2f2; }
        tr:hover { background-color: #e8f4f8; }
        .metadata { background-color: #f0f0f0; padding: 10px; margin-bottom: 20px; }
    </style>
</head>
<body>
    <h1>MACE Database Export</h1>
""")
            
            # Metadata section
            if metadata or True:
                f.write('    <div class="metadata">\n')
                f.write('        <h3>Export Information</h3>\n')
                f.write(f'        <p><strong>Export Date:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>\n')
                f.write(f'        <p><strong>Total Records:</strong> {len(data)}</p>\n')
                if metadata:
                    for key, value in metadata.items():
                        f.write(f'        <p><strong>{key}:</strong> {value}</p>\n')
                f.write('    </div>\n')
                
            # Data table
            f.write('    <table>\n')
            
            # Header
            if data:
                f.write('        <tr>\n')
                for key in data[0].keys():
                    header = key.replace('_', ' ').title()
                    f.write(f'            <th>{self._html_escape(header)}</th>\n')
                f.write('        </tr>\n')
                
                # Data rows
                for record in data:
                    f.write('        <tr>\n')
                    for key in data[0].keys():
                        value = record.get(key, '')
                        if isinstance(value, float):
                            value = f"{value:.6f}".rstrip('0').rstrip('.')
                        f.write(f'            <td>{self._html_escape(str(value))}</td>\n')
                    f.write('        </tr>\n')
                    
            f.write('    </table>\n')
            f.write('</body>\n</html>')
            
        return True
        
    def _latex_escape(self, text: str) -> str:
        """Escape special LaTeX characters."""
        replacements = {
            '&': r'\&',
            '%': r'\%',
            '$': r'\$',
            '#': r'\#',
            '_': r'\_',
            '{': r'\{',
            '}': r'\}',
            '~': r'\textasciitilde{}',
            '^': r'\^{}',
            '\\': r'\textbackslash{}',
        }
        for char, replacement in replacements.items():
            text = text.replace(char, replacement)
        return text
        
    def _html_escape(self, text: str) -> str:
        """Escape HTML special characters."""
        replacements = {
            '&': '&amp;',
            '<': '&lt;',
            '>': '&gt;',
            '"': '&quot;',
            "'": '&#39;',
        }
        for char, replacement in replacements.items():
            text = text.replace(char, replacement)
        return text


def export_materials(db, format: str = 'csv', output_file: str = None,
                    filters: List[str] = None, logic: str = 'AND',
                    properties_only: bool = False,
                    include_properties: List[str] = None,
                    include_structures: bool = False) -> str:
    """
    Convenience function to export materials from database.
    
    Args:
        db: MaterialDatabase instance
        format: Export format
        output_file: Output file path (auto-generated if None)
        filters: Property filters to apply
        logic: Filter logic (AND/OR)
        properties_only: Export properties instead of materials
        include_properties: List of property names to include
        include_structures: Whether to include structure data
        
    Returns:
        Path to exported file
    """
    formatter = ExportFormatter()
    
    # Generate output filename if not provided
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = 'xlsx' if format == 'excel' else format
        output_file = f"mace_export_{timestamp}.{extension}"
        
    # Get data based on options
    if properties_only:
        # Export properties
        data = db.get_all_properties()
        metadata = {'export_type': 'properties'}
    else:
        # Export materials
        if filters:
            # Check if it's a single advanced filter expression
            if len(filters) == 1 and any(op in filters[0] for op in ['(', ')', ' AND ', ' OR ', ' LIKE ', ' IN ', ' IS ']):
                # Use advanced filtering
                data = db.filter_materials_advanced(filters[0])
                metadata = {
                    'export_type': 'filtered_materials',
                    'filters': filters,
                    'filter_type': 'advanced'
                }
            else:
                # Use regular filtering
                data = db.filter_materials_by_properties(filters, logic)
                metadata = {
                    'export_type': 'filtered_materials',
                    'filters': filters,
                    'filter_logic': logic
                }
        else:
            data = db.get_all_materials()
            metadata = {'export_type': 'all_materials'}
            
        # Add properties to materials if requested
        if include_properties and data:
            for material in data:
                mat_id = material['material_id']
                props = db.get_material_properties(mat_id)
                
                # Convert properties to dict
                prop_dict = {}
                for prop in props:
                    prop_name = prop['property_name']
                    if include_properties is None or prop_name in include_properties:
                        prop_dict[prop_name] = prop['property_value']
                        
                material.update(prop_dict)
                
    # Export
    success = formatter.export(
        data,
        output_file,
        format=format,
        properties_filter=include_properties,
        include_structures=include_structures,
        metadata=metadata
    )
    
    return output_file if success else None